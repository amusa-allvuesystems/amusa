#!/usr/bin/env python3
"""Import Net2 'Who's been in today' HTML into RTO Tracker London xlsx."""

import argparse
import re
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter

MONTH_SHEETS = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Aug",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dec",
}


def norm_name(name: str) -> str:
    name = re.sub(r"\s+", " ", name.strip())
    return re.sub(r"\s*,\s*", ", ", name).lower()


def parse_html(path: Path):
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    title = soup.find("h2")
    report_date = None
    if title:
        m = re.search(r"on (\d{1,2} \w+ \d{4})", title.get_text())
        if m:
            report_date = datetime.strptime(m.group(1), "%d %B %Y").date()

    entries = []
    for tr in soup.select("table tr"):
        cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
        if len(cells) == 3 and cells[0] != "Date":
            entries.append(
                {
                    "date": datetime.strptime(cells[0], "%d/%m/%Y").date(),
                    "department": cells[1],
                    "name": cells[2],
                }
            )
    return report_date, entries


def build_date_column_map(ws, year: int, month: int) -> dict[date, str]:
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    mapping = {}
    col = 8  # column I
    day_offset = {"Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4}

    while col < len(row2):
        week_hdr = row1[col] if col < len(row1) else None
        start_day = None
        if isinstance(week_hdr, str):
            m = re.search(r"\((\d+)-", week_hdr)
            if m:
                start_day = int(m.group(1))

        while col < len(row2):
            hdr = row2[col]
            if hdr is None:
                col += 1
                continue
            if "Attn" in str(hdr):
                col += 1
                break
            if hdr in day_offset and start_day is not None:
                d = date(year, month, start_day) + timedelta(days=day_offset[hdr])
                mapping[d] = get_column_letter(col + 1)
                col += 1
            else:
                col += 1
                break
    return mapping


def build_employee_index(ws):
    by_exact = {}
    by_last_first = {}
    for row in ws.iter_rows(min_row=3):
        name = row[1].value
        if not name:
            continue
        n = norm_name(str(name))
        by_exact[n] = row[0].row
        parts = n.split(",", 1)
        if len(parts) == 2:
            key = f"{parts[0].strip()}|{parts[1].strip().split()[0]}"
            by_last_first[key] = row[0].row
    return by_exact, by_last_first


def find_row(name: str, by_exact, by_last_first):
    n = norm_name(name)
    if n in by_exact:
        return by_exact[n]
    parts = n.split(",", 1)
    if len(parts) == 2:
        key = f"{parts[0].strip()}|{parts[1].strip().split()[0]}"
        return by_last_first.get(key)
    return None


def require_file(path: Path, label: str) -> Path:
    if path.is_file():
        return path.resolve()

    cwd = Path.cwd()
    msg = [f"Could not find {label}: {path}"]
    if not path.is_absolute():
        msg.append(f"  (looked in {cwd / path})")

    html_files = sorted(cwd.glob("*.htm")) + sorted(cwd.glob("*.html"))
    if html_files and label == "HTML file":
        msg.append("HTML files in this folder:")
        for f in html_files:
            msg.append(f"  {f.name}")
        msg.append(f'Try: python3 import_net2_attendance.py "{html_files[0].name}" "..." --dry-run')

    xlsx_files = sorted(cwd.glob("*.xlsx"))
    if xlsx_files and label == "Excel file":
        msg.append("Excel files in this folder:")
        for f in xlsx_files:
            msg.append(f"  {f.name}")

    raise SystemExit("\n".join(msg))


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("html", type=Path, help="Net2 HTML export")
    parser.add_argument("xlsx", type=Path, help="RTO Tracker workbook")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    args = parser.parse_args()

    html_path = require_file(args.html, "HTML file")
    xlsx_path = require_file(args.xlsx, "Excel file")

    report_date, entries = parse_html(html_path)
    if not report_date:
        raise SystemExit("Could not determine report date from HTML title.")

    wb = openpyxl.load_workbook(xlsx_path)
    sheet_name = MONTH_SHEETS[report_date.month]
    ws = wb[sheet_name]
    date_cols = build_date_column_map(ws, report_date.year, report_date.month)
    col = date_cols.get(report_date)
    if not col:
        raise SystemExit(f"No column found for {report_date} on sheet {sheet_name}")

    by_exact, by_last_first = build_employee_index(ws)
    updated, skipped, unmatched = [], [], []

    for e in entries:
        if e["department"].startswith("_"):
            skipped.append(e["name"])
            continue
        row = find_row(e["name"], by_exact, by_last_first)
        if not row:
            unmatched.append(e["name"])
            continue
        addr = f"{col}{row}"
        if not args.dry_run:
            ws[addr] = 1
        updated.append(f"{e['name']} -> {sheet_name}!{addr}")

    if not args.dry_run:
        wb.save(xlsx_path)

    print(f"Date: {report_date} -> {sheet_name}!{col}")
    print(f"Marked present: {len(updated)}")
    for line in updated:
        print(f"  {line}")
    if skipped:
        print(f"Skipped (external/ELT): {', '.join(skipped)}")
    if unmatched:
        print(f"No match in tracker: {', '.join(unmatched)}")


if __name__ == "__main__":
    main()
